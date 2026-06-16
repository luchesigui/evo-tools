#!/usr/bin/env python3
"""
Script para envio de mensagens em massa via WhatsApp pelo Kommo CRM.

Autor: EvoTools - Kommo Broadcast
Versão: 1.0.0
"""

import os
import re
import sys
import json
import time
import uuid
import hashlib
import hmac
import base64
import argparse
import logging
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass
from datetime import datetime
from email.utils import formatdate

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook


# Configurações globais
MAX_RETRIES = 3
RATE_LIMIT_DELAY = 1.0  # Delay entre requests em segundos
DEFAULT_DELAY = 2.0  # Delay padrão entre mensagens


@dataclass
class Contact:
    """Representa um contato para broadcast."""
    nome: str
    telefone: str
    telefone_sanitizado: str
    kommo_contact: Optional[Dict[str, Any]] = None


@dataclass
class BroadcastStats:
    """Estatísticas do broadcast."""
    total_planilha: int = 0
    encontrados_kommo: int = 0
    nao_encontrados: int = 0
    enviados: int = 0
    erros: int = 0


class KommoAPI:
    """Cliente para interagir com a API do Kommo CRM."""
    
    def __init__(self, domain: str, client_id: str, client_secret: str, 
                 access_token: str, refresh_token: str, dry_run: bool = False):
        self.domain = domain
        self.client_id = client_id
        self.client_secret = client_secret
        self.access_token = access_token
        self.refresh_token = refresh_token
        self.dry_run = dry_run
        self.base_url = f"https://{domain}"
        self.session = requests.Session()
        
        # IDs importantes que serão descobertos durante a inicialização
        self.phone_field_id = None
        self.account_id = None
        
        self.logger = logging.getLogger(__name__)
    
    def _make_request(self, method: str, endpoint: str, **kwargs) -> requests.Response:
        """Faz uma requisição HTTP com tratamento de erros e rate limiting."""
        url = f"{self.base_url}{endpoint}"
        
        if self.dry_run and method.upper() in ['POST', 'PATCH', 'PUT', 'DELETE']:
            self.logger.info(f"[DRY RUN] {method.upper()} {url}")
            if 'json' in kwargs:
                self.logger.info(f"[DRY RUN] Body: {json.dumps(kwargs['json'], indent=2)}")
            # Retorna uma resposta mock para dry run
            mock_response = requests.Response()
            mock_response.status_code = 200
            mock_response._content = json.dumps({"_embedded": {"contacts": []}}).encode()
            return mock_response
        
        headers = kwargs.get('headers', {})
        headers.update({
            'Authorization': f'Bearer {self.access_token}',
            'Content-Type': 'application/json'
        })
        kwargs['headers'] = headers
        
        for attempt in range(MAX_RETRIES):
            try:
                self.logger.debug(f"Fazendo request {method.upper()} para {url}")
                response = self.session.request(method, url, **kwargs)
                
                # Rate limiting
                if response.status_code == 429:
                    wait_time = RATE_LIMIT_DELAY * (2 ** attempt)
                    self.logger.warning(f"Rate limit atingido. Aguardando {wait_time}s...")
                    time.sleep(wait_time)
                    continue
                
                # Token expirado
                if response.status_code == 401:
                    self.logger.info("Token expirado. Tentando refresh...")
                    if self.refresh_access_token():
                        headers['Authorization'] = f'Bearer {self.access_token}'
                        kwargs['headers'] = headers
                        continue
                    else:
                        raise Exception("Não foi possível renovar o token")
                
                if response.status_code >= 400:
                    self.logger.error(f"Resposta de erro da API: {response.text}")
                response.raise_for_status()
                time.sleep(RATE_LIMIT_DELAY)  # Rate limiting preventivo
                return response
                
            except requests.exceptions.RequestException as e:
                self.logger.warning(f"Erro na tentativa {attempt + 1}: {e}")
                if attempt == MAX_RETRIES - 1:
                    raise
                time.sleep(RATE_LIMIT_DELAY * (attempt + 1))
        
        raise Exception(f"Falha após {MAX_RETRIES} tentativas")
    
    def refresh_access_token(self) -> bool:
        """Renova o access token usando o refresh token."""
        if not self.refresh_token:
            self.logger.warning("Nenhum refresh token disponível para renovação.")
            return False
        try:
            data = {
                'client_id': self.client_id,
                'client_secret': self.client_secret,
                'grant_type': 'refresh_token',
                'refresh_token': self.refresh_token
            }
            
            response = requests.post(
                f"{self.base_url}/oauth2/access_token",
                data=data,
                headers={'Content-Type': 'application/x-www-form-urlencoded'}
            )
            response.raise_for_status()
            
            token_data = response.json()
            self.access_token = token_data['access_token']
            self.refresh_token = token_data['refresh_token']
            
            self.logger.info("Token renovado com sucesso")
            return True
            
        except Exception as e:
            self.logger.error(f"Erro ao renovar token: {e}")
            return False
    
    def get_custom_fields(self) -> List[Dict[str, Any]]:
        """Busca todos os campos customizados de contatos."""
        response = self._make_request('GET', '/api/v4/contacts/custom_fields')
        data = response.json()
        return data.get('_embedded', {}).get('custom_fields', [])
    
    def find_phone_field_id(self) -> Optional[int]:
        """Encontra o field_id do campo de telefone padrão."""
        custom_fields = self.get_custom_fields()
        
        # Procura por campos de telefone
        for field in custom_fields:
            if field.get('type') == 'multitext' and 'phone' in field.get('code', '').lower():
                return field['id']
        
        return None
    
    def get_account_info(self) -> Dict[str, Any]:
        """Busca informações da conta."""
        response = self._make_request('GET', '/api/v4/account')
        return response.json()
    
    def search_contact_by_phone(self, phone: str) -> Optional[Dict[str, Any]]:
        """Busca um contato pelo telefone."""
        params = {'query': phone, 'limit': 1}
        response = self._make_request('GET', '/api/v4/contacts', params=params)
        data = response.json()
        
        contacts = data.get('_embedded', {}).get('contacts', [])
        return contacts[0] if contacts else None
    
    def initialize(self):
        """Inicializa IDs necessários para o broadcast."""
        self.logger.info("Inicializando API do Kommo...")
        
        # Busca field_id do telefone
        self.phone_field_id = self.find_phone_field_id()
        if self.phone_field_id:
            self.logger.info(f"Campo de telefone encontrado com ID: {self.phone_field_id}")
        else:
            self.logger.warning("Campo de telefone não encontrado - usando busca genérica")
        
        # Busca informações da conta
        try:
            account_info = self.get_account_info()
            self.account_id = account_info.get('id')
            if self.account_id:
                self.logger.info(f"Account ID obtido: {self.account_id}")
            else:
                self.logger.warning("Account ID não encontrado")
        except Exception as e:
            self.logger.error(f"Erro ao obter informações da conta: {e}")


class KommoChatsAPI:
    """Cliente para interagir com a Chats API do Kommo (amojo)."""
    
    def __init__(self, channel_id: str, channel_secret: str, scope_id: str, 
                 account_id: str, dry_run: bool = False):
        self.channel_id = channel_id
        self.channel_secret = channel_secret
        self.scope_id = scope_id
        self.account_id = account_id
        self.dry_run = dry_run
        self.base_url = "https://amojo.kommo.com"
        self.session = requests.Session()
        
        self.logger = logging.getLogger(__name__)
    
    def _generate_signature(self, method: str, content_md5: str, content_type: str, 
                          date: str, path: str) -> str:
        """Gera a assinatura HMAC-SHA1 para a Chats API."""
        message = f"{method.upper()}\n{content_md5}\n{content_type}\n{date}\n{path}"
        signature = hmac.new(
            self.channel_secret.encode(),
            message.encode(),
            hashlib.sha1
        ).digest()
        return base64.b64encode(signature).decode()
    
    def _make_chats_request(self, method: str, path: str, data: Dict[str, Any]) -> requests.Response:
        """Faz uma requisição para a Chats API com assinatura HMAC."""
        url = f"{self.base_url}{path}"
        body = json.dumps(data, separators=(',', ':'))
        
        if self.dry_run:
            self.logger.info(f"[DRY RUN] {method.upper()} {url}")
            self.logger.info(f"[DRY RUN] Body: {json.dumps(data, indent=2)}")
            # Retorna uma resposta mock para dry run
            mock_response = requests.Response()
            mock_response.status_code = 200
            mock_response._content = json.dumps({"result": "success"}).encode()
            return mock_response
        
        # Calcula hashes e assinatura
        content_type = "application/json"
        content_md5 = hashlib.md5(body.encode()).hexdigest()
        date = formatdate(timeval=None, localtime=False, usegmt=True)
        
        signature = self._generate_signature(method, content_md5, content_type, date, path)
        
        headers = {
            'Date': date,
            'Content-Type': content_type,
            'Content-MD5': content_md5,
            'X-Signature': signature
        }
        
        for attempt in range(MAX_RETRIES):
            try:
                self.logger.debug(f"Fazendo request {method.upper()} para {url}")
                response = self.session.request(method, url, data=body, headers=headers)
                
                # Rate limiting
                if response.status_code == 429:
                    wait_time = RATE_LIMIT_DELAY * (2 ** attempt)
                    self.logger.warning(f"Rate limit atingido na Chats API. Aguardando {wait_time}s...")
                    time.sleep(wait_time)
                    continue
                
                if response.status_code >= 400:
                    self.logger.error(f"Erro da Chats API: {response.text}")
                response.raise_for_status()
                time.sleep(RATE_LIMIT_DELAY)  # Rate limiting preventivo
                return response
                
            except requests.exceptions.RequestException as e:
                self.logger.warning(f"Erro na tentativa {attempt + 1} da Chats API: {e}")
                if attempt == MAX_RETRIES - 1:
                    raise
                time.sleep(RATE_LIMIT_DELAY * (attempt + 1))
        
        raise Exception(f"Falha na Chats API após {MAX_RETRIES} tentativas")
    
    def create_chat(self, phone: str, name: str) -> Dict[str, Any]:
        """Cria um chat para um contato."""
        path = f"/v2/origin/custom/{self.scope_id}/chats"
        data = {
            "conversation_id": phone,
            "user": {
                "id": phone,
                "name": name,
                "profile": {
                    "phone": phone
                }
            }
        }
        
        response = self._make_chats_request('POST', path, data)
        return response.json()
    
    def send_message(self, conversation_id: str, message_text: str, sender_name: str) -> Dict[str, Any]:
        """Envia uma mensagem para um chat."""
        path = f"/v2/origin/custom/{self.scope_id}"
        
        data = {
            "account_id": self.account_id,
            "event_type": "new_message",
            "payload": {
                "timestamp": int(time.time()),
                "msgid": str(uuid.uuid4()),
                "conversation_id": conversation_id,
                "sender": {
                    "id": "system",
                    "name": sender_name or "Sistema"
                },
                "message": {
                    "type": "text",
                    "text": message_text
                },
                "silent": True
            }
        }
        
        response = self._make_chats_request('POST', path, data)
        return response.json()


def sanitize_phone(phone: str) -> str:
    """Remove tudo que não for dígito do telefone."""
    return re.sub(r'\D', '', str(phone))


def read_xlsx_contacts(file_path: str) -> List[Contact]:
    """Lê contatos de um arquivo XLSX."""
    logger = logging.getLogger(__name__)
    
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")
    
    logger.info(f"Lendo arquivo XLSX: {file_path}")
    
    try:
        workbook = load_workbook(file_path, read_only=True)
        sheet = workbook.active
        
        contacts = []
        headers_row = 1
        
        # Identifica as colunas
        headers = {}
        for col_idx, cell in enumerate(sheet[headers_row], 1):
            if cell and cell.value:
                header_lower = str(cell.value).lower().strip()
                if 'nome' in header_lower or 'name' in header_lower:
                    headers['nome'] = col_idx
                elif 'telefone' in header_lower or 'phone' in header_lower or 'fone' in header_lower:
                    headers['telefone'] = col_idx
        
        if 'telefone' not in headers:
            # Se não encontrou headers, assume colunas padrão: A=nome, B=telefone
            headers = {'nome': 1, 'telefone': 2}
            headers_row = 0  # Começa da primeira linha
            logger.warning("Headers não identificados. Assumindo: A=nome, B=telefone")
        
        logger.info(f"Colunas identificadas: {headers}")
        
        # Lê os dados
        if sheet is None:
            raise Exception("Planilha não encontrada")
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=headers_row + 1), headers_row + 1):
            try:
                nome_cell = row[headers.get('nome', 1) - 1] if headers.get('nome') else None
                telefone_cell = row[headers.get('telefone', 2) - 1] if headers.get('telefone') else None
                
                nome = str(nome_cell.value).strip() if nome_cell and nome_cell.value else f"Contato {row_idx}"
                telefone_raw = str(telefone_cell.value).strip() if telefone_cell and telefone_cell.value else ""
                
                if not telefone_raw or telefone_raw.lower() in ['none', 'null', '']:
                    continue
                
                telefone_sanitizado = sanitize_phone(telefone_raw)
                
                if len(telefone_sanitizado) < 10:
                    logger.warning(f"Linha {row_idx}: Telefone muito curto ignorado: {telefone_raw}")
                    continue
                
                contact = Contact(
                    nome=nome,
                    telefone=telefone_raw,
                    telefone_sanitizado=telefone_sanitizado
                )
                contacts.append(contact)
                
            except Exception as e:
                logger.error(f"Erro ao processar linha {row_idx}: {e}")
        
        logger.info(f"Total de {len(contacts)} contatos carregados do XLSX")
        return contacts
        
    except Exception as e:
        logger.error(f"Erro ao ler arquivo XLSX: {e}")
        raise


def search_contacts_in_kommo(contacts: List[Contact], kommo_api: KommoAPI) -> Tuple[List[Contact], List[Contact]]:
    """Busca contatos no Kommo e separa em encontrados/não encontrados."""
    logger = logging.getLogger(__name__)
    
    found_contacts = []
    not_found_contacts = []
    
    total = len(contacts)
    
    for i, contact in enumerate(contacts, 1):
        logger.info(f"Buscando contato {i}/{total}: {contact.nome} ({contact.telefone_sanitizado})")
        
        try:
            kommo_contact = kommo_api.search_contact_by_phone(contact.telefone_sanitizado)
            
            if kommo_contact:
                contact.kommo_contact = kommo_contact
                found_contacts.append(contact)
                logger.debug(f"✓ Contato encontrado no Kommo: {contact.nome}")
            else:
                not_found_contacts.append(contact)
                logger.debug(f"✗ Contato não encontrado no Kommo: {contact.nome}")
        
        except Exception as e:
            logger.error(f"Erro ao buscar contato {contact.nome}: {e}")
            not_found_contacts.append(contact)
    
    logger.info(f"Busca finalizada: {len(found_contacts)} encontrados, {len(not_found_contacts)} não encontrados")
    return found_contacts, not_found_contacts


def send_broadcasts(contacts: List[Contact], chats_api: KommoChatsAPI, 
                   message_template: str, delay: float = DEFAULT_DELAY) -> BroadcastStats:
    """Envia mensagens para os contatos encontrados."""
    logger = logging.getLogger(__name__)
    
    stats = BroadcastStats()
    total = len(contacts)
    
    logger.info(f"Iniciando envio para {total} contatos...")
    
    for i, contact in enumerate(contacts, 1):
        try:
            # Substitui variáveis na mensagem
            message_text = message_template.format(nome=contact.nome)
            
            logger.info(f"Enviando {i}/{total} para {contact.nome} ({contact.telefone_sanitizado})")
            
            # Cria o chat
            chat_result = chats_api.create_chat(contact.telefone_sanitizado, contact.nome)
            logger.debug(f"Chat criado: {chat_result}")
            
            # Envia a mensagem
            message_result = chats_api.send_message(
                contact.telefone_sanitizado,
                message_text,
                "EvoTools Broadcast"
            )
            logger.debug(f"Mensagem enviada: {message_result}")
            
            stats.enviados += 1
            logger.info(f"✓ Mensagem enviada com sucesso para {contact.nome}")
            
            # Rate limiting
            if i < total:  # Não faz delay no último
                logger.debug(f"Aguardando {delay}s antes da próxima mensagem...")
                time.sleep(delay)
        
        except Exception as e:
            stats.erros += 1
            logger.error(f"✗ Erro ao enviar mensagem para {contact.nome}: {e}")
    
    logger.info(f"Envio finalizado: {stats.enviados} enviados, {stats.erros} erros")
    return stats


def setup_logging(verbose: bool = False):
    """Configura o logging."""
    level = logging.DEBUG if verbose else logging.INFO
    format_str = '%(asctime)s - %(levelname)s - %(message)s'
    
    logging.basicConfig(
        level=level,
        format=format_str,
        handlers=[
            logging.StreamHandler(sys.stdout)
        ]
    )


def main():
    """Função principal."""
    parser = argparse.ArgumentParser(description='Envio de mensagens em massa via WhatsApp pelo Kommo CRM')
    parser.add_argument('--dry-run', action='store_true', help='Simular sem enviar mensagens')
    parser.add_argument('--verbose', '-v', action='store_true', help='Logs detalhados')
    parser.add_argument('--message', '-m', help='Texto da mensagem (sobrescreve .env)')
    parser.add_argument('--xlsx', help='Arquivo XLSX com contatos (sobrescreve .env)')
    parser.add_argument('--delay', type=float, default=DEFAULT_DELAY, 
                       help=f'Delay entre mensagens em segundos (padrão: {DEFAULT_DELAY})')
    
    args = parser.parse_args()
    
    # Configura logging
    setup_logging(args.verbose)
    logger = logging.getLogger(__name__)
    
    # Carrega variáveis de ambiente
    load_dotenv()
    
    # Valida variáveis obrigatórias
    required_vars = [
        'KOMMO_DOMAIN', 'KOMMO_CLIENT_ID', 'KOMMO_CLIENT_SECRET',
        'KOMMO_ACCESS_TOKEN', 'KOMMO_REFRESH_TOKEN',
        'KOMMO_CHANNEL_ID', 'KOMMO_CHANNEL_SECRET', 'KOMMO_SCOPE_ID'
    ]
    
    missing_vars = [var for var in required_vars if not os.getenv(var)]
    if missing_vars:
        logger.error(f"Variáveis de ambiente obrigatórias não encontradas: {', '.join(missing_vars)}")
        logger.error("Copie .env.example para .env e configure as credenciais")
        sys.exit(1)
    
    # Parâmetros
    xlsx_file = args.xlsx or os.getenv('XLSX_FILE_PATH')
    message_template = args.message or os.getenv('MESSAGE_TEXT', 'Olá {nome}!')
    
    if not xlsx_file:
        logger.error("Arquivo XLSX não especificado. Use --xlsx ou configure XLSX_FILE_PATH no .env")
        sys.exit(1)
    
    if args.dry_run:
        logger.info("=== MODO DRY RUN ATIVADO - NENHUMA MENSAGEM SERÁ ENVIADA ===")
    
    logger.info("Iniciando Kommo Broadcast...")
    logger.info(f"Arquivo XLSX: {xlsx_file}")
    logger.info(f"Template de mensagem: {message_template}")
    logger.info(f"Delay entre mensagens: {args.delay}s")
    
    try:
        # 1. Lê contatos do XLSX
        logger.info("1/4 - Lendo contatos do arquivo XLSX...")
        contacts = read_xlsx_contacts(xlsx_file)
        
        if not contacts:
            logger.error("Nenhum contato válido encontrado no arquivo XLSX")
            sys.exit(1)
        
        # 2. Inicializa Kommo API
        logger.info("2/4 - Inicializando Kommo API...")
        
        # Valida que temos todas as credenciais necessárias
        domain = os.getenv('KOMMO_DOMAIN')
        client_id = os.getenv('KOMMO_CLIENT_ID')
        client_secret = os.getenv('KOMMO_CLIENT_SECRET')
        access_token = os.getenv('KOMMO_ACCESS_TOKEN')
        refresh_token = os.getenv('KOMMO_REFRESH_TOKEN')
        
        if not all([domain, client_id, client_secret, access_token, refresh_token]):
            logger.error("Credenciais do Kommo incompletas")
            sys.exit(1)
        
        # Assertions para satisfazer o type checker
        assert domain is not None
        assert client_id is not None
        assert client_secret is not None
        assert access_token is not None
        assert refresh_token is not None
        
        kommo_api = KommoAPI(
            domain=domain,
            client_id=client_id,
            client_secret=client_secret,
            access_token=access_token,
            refresh_token=refresh_token,
            dry_run=args.dry_run
        )
        kommo_api.initialize()
        
        # 3. Busca contatos no Kommo
        logger.info("3/4 - Buscando contatos no Kommo...")
        found_contacts, not_found_contacts = search_contacts_in_kommo(contacts, kommo_api)
        
        # 4. Envia mensagens
        stats = BroadcastStats()
        stats.total_planilha = len(contacts)
        stats.encontrados_kommo = len(found_contacts)
        stats.nao_encontrados = len(not_found_contacts)
        
        if found_contacts:
            logger.info("4/4 - Enviando mensagens...")
            
            # Inicializa Chats API
            channel_id = os.getenv('KOMMO_CHANNEL_ID')
            channel_secret = os.getenv('KOMMO_CHANNEL_SECRET')
            scope_id = os.getenv('KOMMO_SCOPE_ID')
            
            if not all([channel_id, channel_secret, scope_id, kommo_api.account_id]):
                logger.error("Credenciais da Chats API incompletas")
                sys.exit(1)
            
            # Assertions para satisfazer o type checker
            assert channel_id is not None
            assert channel_secret is not None
            assert scope_id is not None
            assert kommo_api.account_id is not None
            
            chats_api = KommoChatsAPI(
                channel_id=channel_id,
                channel_secret=channel_secret,
                scope_id=scope_id,
                account_id=kommo_api.account_id,
                dry_run=args.dry_run
            )
            
            broadcast_stats = send_broadcasts(found_contacts, chats_api, message_template, args.delay)
            stats.enviados = broadcast_stats.enviados
            stats.erros = broadcast_stats.erros
        else:
            logger.warning("Nenhum contato encontrado no Kommo para envio")
        
        # Relatório final
        logger.info("\n" + "="*60)
        logger.info("RELATÓRIO FINAL DO BROADCAST")
        logger.info("="*60)
        logger.info(f"Total na planilha:      {stats.total_planilha}")
        logger.info(f"Encontrados no Kommo:   {stats.encontrados_kommo}")
        logger.info(f"Não encontrados:        {stats.nao_encontrados}")
        logger.info(f"Mensagens enviadas:     {stats.enviados}")
        logger.info(f"Erros no envio:         {stats.erros}")
        
        if args.dry_run:
            logger.info("\n⚠️  MODO DRY RUN - Nenhuma mensagem foi realmente enviada")
        
        # Lista contatos não encontrados
        if not_found_contacts and args.verbose:
            logger.info(f"\nContatos não encontrados no Kommo ({len(not_found_contacts)}):")
            for contact in not_found_contacts[:10]:  # Mostra apenas os primeiros 10
                logger.info(f"  - {contact.nome} ({contact.telefone_sanitizado})")
            if len(not_found_contacts) > 10:
                logger.info(f"  ... e mais {len(not_found_contacts) - 10}")
        
        logger.info("="*60)
        logger.info("Broadcast finalizado!")
        
    except KeyboardInterrupt:
        logger.warning("\nBroadcast interrompido pelo usuário")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Erro durante o broadcast: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()